import sys
import os

# Add the root directory to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tkinter import messagebox

from GUI.MainFormGUI import worksheet, worksheet2, worksheet3, worksheet4, worksheet5

#Xuất Excel
import ezsheets
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime 

class Excel_Create:
    def XuatExcel5(self):
        # Function to get data from a specified range
        def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(sheet[row, col])
                data.append(row_data)
            return data

        # Download the specific Google Sheet
        ss = ezsheets.Spreadsheet("1tTAZapKjFJ21FYJGoEZBaIYRmHWv2LmW_G4lwZ2pOUE")

        # Specify the sheet, columns, and rows
        # lag
        sheet_name = 'sheet 4'  # Change this to the specific sheet name
        start_row = 1  # Skip the header row
        end_row = 6
        start_col = 3
        test = worksheet4.get_all_values()
        end_col = len([row[1] for row in test] )

        sheet = ss[sheet_name]
        data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

        # Create a new Excel file and write data to it
        wb = Workbook()
        ws = wb.active

        # Write headers and data to the new Excel sheet
        headers = [
            'ID',	'Code',	'Full name',	'Phone',	'Main class',	'Review class'																																								
        ]

        # Define styles
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # Write the title and apply styles
        title_cell = ws.cell(row=1, column=1, value="Quản Lý Lớp Ôn")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = header_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        # Write the headers and apply styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write the data and apply borders
        for col_idx, col_data in enumerate(data, start=1):
            for row_idx, value in enumerate(col_data, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths to fit the content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate unique file name with date and time
        current_time = datetime.now().strftime("%H-%M-%S-%d-%m-%Y")
        file_path = f'D:\\QLO-{current_time}.xlsx'

        # Save the new Excel file
        wb.save(file_path)
        messagebox.showinfo("Success", "Download the file successfully, please check your D drive!")



    def XuatExcel3(self):
        # Function to get data from a specified range
        def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(sheet[row, col])
                data.append(row_data)
            return data

        # Download the specific Google Sheet
        ss = ezsheets.Spreadsheet("1tTAZapKjFJ21FYJGoEZBaIYRmHWv2LmW_G4lwZ2pOUE")

        # Specify the sheet, columns, and rows
        # lag
        sheet_name = 'sheet 1'  # Change this to the specific sheet name
        start_row = 1  # Skip the header row
        end_row = 9
        start_col = 3
        test = worksheet.get_all_values()
        end_col = len([row[1] for row in test] )

        sheet = ss[sheet_name]
        data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

        # Create a new Excel file and write data to it
        wb = Workbook()
        ws = wb.active

        # Write headers and data to the new Excel sheet
        headers = [
            'ID',	'Main class',	'Book 1',	'Book 2',	'Book 3',	'Book 4',	'Book 5',	'Main teacher',	'FOREIGN TEACHER'																																									
        ]

        # Define styles
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # Write the title and apply styles
        title_cell = ws.cell(row=1, column=1, value="Quản Lý Sách")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = header_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        # Write the headers and apply styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write the data and apply borders
        for col_idx, col_data in enumerate(data, start=1):
            for row_idx, value in enumerate(col_data, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths to fit the content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate unique file name with date and time
        current_time = datetime.now().strftime("%H-%M-%S-%d-%m-%Y")
        file_path = f'D:\\QLS-{current_time}.xlsx'

        # Save the new Excel file
        wb.save(file_path)
        messagebox.showinfo("Success", "Download the file successfully, please check your D drive!")

    def XuatExcel(self):
        # Function to get data from a specified range
        def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(sheet[row, col])
                data.append(row_data)
            return data

        # Download the specific Google Sheet
        ss = ezsheets.Spreadsheet("1tTAZapKjFJ21FYJGoEZBaIYRmHWv2LmW_G4lwZ2pOUE")

        # Specify the sheet, columns, and rows
        # lag
        sheet_name = 'sheet 3'  # Change this to the specific sheet name
        start_row = 1  # Skip the header row
        end_row = 7
        start_col = 3
        test = worksheet3.get_all_values()
        end_col = len([row[1] for row in test] )

        sheet = ss[sheet_name]
        data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

        # Create a new Excel file and write data to it
        wb = Workbook()
        ws = wb.active

        # Write headers and data to the new Excel sheet
        headers = [
            "CLASSNO", "MAIN CLASS", "STUDYING DAY", "STUDYING TIME", "ROOM", "TEACHER", "FOREIGN TEACHER"
        ]

        # Define styles
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # Write the title and apply styles
        title_cell = ws.cell(row=1, column=1, value="Quản Lý Lớp Học")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = header_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

        # Write the headers and apply styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write the data and apply borders
        for col_idx, col_data in enumerate(data, start=1):
            for row_idx, value in enumerate(col_data, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths to fit the content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate unique file name with date and time
        current_time = datetime.now().strftime("%H-%M-%S-%d-%m-%Y")
        file_path = f'D:\\QLLH-{current_time}.xlsx'

        # Save the new Excel file
        wb.save(file_path)
        messagebox.showinfo("Success", "Download the file successfully, please check your D drive!")

    def XuatExcel12(self):
        # Function to get data from a specified range
        def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(sheet[row, col])
                data.append(row_data)
            return data

        # Download the specific Google Sheet
        ss = ezsheets.Spreadsheet("1tTAZapKjFJ21FYJGoEZBaIYRmHWv2LmW_G4lwZ2pOUE")

        # Specify the sheet, columns, and rows
        # lag
        sheet_name = 'sheet 2'  # Change this to the specific sheet name
        start_row = 1  # Skip the header row
        end_row = 50
        start_col = 3
        test = worksheet2.get_all_values()
        end_col = len([row[1] for row in test] )

        sheet = ss[sheet_name]
        data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

        # Create a new Excel file and write data to it
        wb = Workbook()
        ws = wb.active

        # Write headers and data to the new Excel sheet
        headers = [
            

            "ID",	"FULL NAME",	"BIRTHDAY (DOB)",	"MAIN CLASS",	"CURRENT LEVEL",	'STUDYING DAY',	'STUDYING TIME',	'TEL',	'ADDRESS',	'PARENT NAME',	'ENROLCAMP',	
            'MAIN CAMP',	'TOTAL FEE',	'MAIN FEE',	'NEW COMER',	'STARTING OFF MONTH',	'STARTING QUIT MONTH',	'CERTIFICATE',	'PUBLIC SCHOOL',	
            'SUB TEL',	'STARTING TRANSFER MONTH',	'TEACHER',	'EXAM DAY',	'EXAM TIME',	'EXAM INVIGILATOR',	'LISTENING 1',	'SPEAKING 1',	
            'READING & WRITING 1',	'TOTAL GRADE 1',	'PERCENT 1',	'LISTENING 2',	'SPEAKING 2',	'READING & WRITING 2',	'TOTAL GRADE 2',	
            'PERCENT 2',	'LISTENING 3',	'SPEAKING 3',	'READING & WRITING 3',	'TOTAL GRADE 3',	'PERCENT 3',	'LISTENING 4',	'SPEAKING 4',	
            'READING & WRITING 4',	'TOTAL GRADE 4',	'PERCENT 4',	'LISTENING 5',	'SPEAKING 5',	'READING & WRITING 5',	'TOTAL GRADE 5',	
            'PERCENT 5'
        ]

        # Define styles
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # Write the title and apply styles
        title_cell = ws.cell(row=1, column=1, value="Quản Lý Lớp Học")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = header_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=50)

        # Write the headers and apply styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write the data and apply borders
        for col_idx, col_data in enumerate(data, start=1):
            for row_idx, value in enumerate(col_data, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths to fit the content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate unique file name with date and time
        current_time = datetime.now().strftime("%H-%M-%S-%d-%m-%Y")
        file_path = f'D:\\QLHS-{current_time}.xlsx'

        # Save the new Excel file
        wb.save(file_path)
        messagebox.showinfo("Success", "Download the file successfully, please check your D drive!")
    
    def XuatExcel4(self):
        # Function to get data from a specified range
        def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(sheet[row, col])
                data.append(row_data)
            return data

        # Download the specific Google Sheet
        ss = ezsheets.Spreadsheet("1tTAZapKjFJ21FYJGoEZBaIYRmHWv2LmW_G4lwZ2pOUE")

        # Specify the sheet, columns, and rows
        # lag
        sheet_name = 'sheet 5'  # Change this to the specific sheet name
        start_row = 1  # Skip the header row
        end_row = 8
        start_col = 3
        test = worksheet5.get_all_values()
        end_col = len([row[1] for row in test] )

        sheet = ss[sheet_name]
        data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

        # Create a new Excel file and write data to it
        wb = Workbook()
        ws = wb.active

        # Write headers and data to the new Excel sheet
        headers = [
            'ID',	'Code',	'Full name',	'Phone',	'Main class',	'Class change',	'Reason for changing class',	'Start date'
        ]

        # Define styles
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        title_font = Font(bold=True, size=14)

        # Write the title and apply styles
        title_cell = ws.cell(row=1, column=1, value="Quản Lý Chuyển Lớp")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = header_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        # Write the headers and apply styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write the data and apply borders
        for col_idx, col_data in enumerate(data, start=1):
            for row_idx, value in enumerate(col_data, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths to fit the content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate unique file name with date and time
        current_time = datetime.now().strftime("%H-%M-%S-%d-%m-%Y")
        file_path = f'D:\\QLCL-{current_time}.xlsx'

        # Save the new Excel file
        wb.save(file_path)
        messagebox.showinfo("Success", "Download the file successfully, please check your D drive!")