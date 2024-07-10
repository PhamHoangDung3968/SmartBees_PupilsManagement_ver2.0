# import gspread

# gs = gspread.service_account("cre.json")
# sht = gs.open_by_key("1RPL8Tv_JctB7icajUTBoEq1lMO8XYb3sxySdGHJGgvY")
# value = sht.sheet1.acell("B1").value
# # print(value)

# worksheet = sht.sheet1

# # Lấy dữ liệu từ ô A7 đến H15
# cell_range = "A7:H15"
# values = worksheet.get(cell_range)

# # In kết quả
# for row in values:
#     print(row)





# import gspread

# gs = gspread.service_account("cre.json")  # Assuming your credentials file is named "cre.json"
# sht = gs.open_by_key("1RPL8Tv_JctB7icajUTBoEq1lMO8XYb3sxySdGHJGgvY")
# worksheet = sht.sheet1

# # Get all values using `get_all_values()`
# all_values = worksheet.get_all_values()
# # Print the results
# for row in all_values:
#     print(row[:3])


# import gspread

# # Giả sử tệp chứng chỉ của bạn có tên là "cre.json"
# gs = gspread.service_account("cre.json")

# # Mở bảng tính bằng cách sử dụng khóa của nó
# sht = gs.open_by_key("1RPL8Tv_JctB7icajUTBoEq1lMO8XYb3sxySdGHJGgvY")

# # Lấy bảng tính đầu tiên (Sheet1)
# worksheet = sht.sheet1

# # Lấy tất cả giá trị từ cột A đến C
# values_list = worksheet.get_all_values()
# result_list = [row[:5] for row in values_list]

'''
import ezsheets
ss = ezsheets.Spreadsheet("1RPL8Tv_JctB7icajUTBoEq1lMO8XYb3sxySdGHJGgvY")
ss.downloadAsExcel()
'''

import ezsheets
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Function to get data from a specified range
def get_data_from_range(sheet, start_row, end_row, start_col, end_col):
    data = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            row_data.append(sheet[row, col])
        data.append(row_data)
    return data

# Download the specific Google Sheet
ss = ezsheets.Spreadsheet("1RPL8Tv_JctB7icajUTBoEq1lMO8XYb3sxySdGHJGgvY")

# Specify the sheet, columns, and rows
sheet_name = 'sheet 1'  # Change this to the specific sheet name
start_row = 1
end_row = 17
start_col = 1
end_col = 13

sheet = ss[sheet_name]
data = get_data_from_range(sheet, start_row, end_row, start_col, end_col)

# Create a new Excel file and write data to it
wb = Workbook()
ws = wb.active

# Write headers and data to the new Excel sheet
headers = [
    "ID_BOOK", "CAMBRIDGE LEVEL", "PROGRESS", "MAIN BOOK", "SKILL BOOK 1",
    "VOCAB BOOK", "SKILL BOOK 2", "SKILL BOOK 3", "SKILL BOOK 4",
    "GRAMMAR BOOK", "TEST BOOK", "VIDEOS-MOVIES", "PICTURES-CARDS"
]

# Define styles
header_font = Font(bold=True, color="000000")
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)
title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
title_font = Font(bold=True, size=14)

# Write the title and apply styles
title_cell = ws.cell(row=1, column=1, value="Quản Lý Sách")
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = header_alignment
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)

# Write the headers and apply styles
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Write the data and apply borders
for row_idx, row_data in enumerate(data, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Adjust column widths to fit the content
for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    max_length = 0
    column = get_column_letter(col[0].column)  # Get the column name
    for cell in col:
        if cell.value is not None:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the new Excel file
file_path = 'new_file.xlsx'
wb.save(file_path)

print(f"Data copied to {file_path}")

